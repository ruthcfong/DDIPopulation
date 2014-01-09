import sys

# for parsing XML
from lxml import etree

# for handling old Excel files with the .xls extension
import xlrd

# for handling Excel 2007+ .xlsx files
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException

DEBUG = True

REQUIRED_NUM_ARGS = 2
OPTIONAL_NUM_ARGS = 1
EXCEL_ARG_INDEX = 1
XML_ARG_INDEX = 2
OUTPUT_XML_ARG_INDEX = 3

VAR_TAG = "var"
DESCRIPTION_TAG = "txt"
QUESTION_TAG = "qstn"
LITERAL_QUESTION_TAG = "qstnLit"

SEARCH_FULL_TREE = ".//"
IGNORE_NAMESPACE = "{*}"

OUTPUT_XML_PREFIX = "ddi_"

OLD_EXCEL_EXTENSION = "xls"

SURVEY_SHEET_NAME = "survey"

DESCRIPTION_COLUMN_NAME = "name"
LITERAL_QUESTION_COLUMN_NAME = "label::English"

NULL_INDEX = -1

EXAMPLE_FILES = "example_files/"
DEBUG_XML_NAME = "testing.xml"
DEBUG_XML = EXAMPLE_FILES + DEBUG_XML_NAME
DEBUG_OUTPUT_XML_NAME = OUTPUT_XML_PREFIX + DEBUG_XML_NAME
DEBUG_EXCEL_FILE_NAME = "example.xlsx"
DEBUG_EXCEL = EXAMPLE_FILES + DEBUG_EXCEL_FILE_NAME

def open_xls_as_xlsx(filename):
    '''
    Returns a openpyxl.workbook object for a .xls file
    http://stackoverflow.com/questions/9918646/how-to-convert-xls-to-xlsx
    '''
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.get_active_sheet()

    for row in xrange(0, nrows):
        for col in xrange(0, ncols):
            sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)

    return book1

def main():
    '''
    TODO for version 1:
    - Remove DEBUG variable (or turn it off)

    TODO for version 2:
    - add command-line option to universe
    - add --help flag
    - make command-line arguments prettier
    '''

    # handle command line arguments
    # TODO: Add --help, -h flag
    # TODO: Handle command line arguments (2; third is optional for output file)
    if DEBUG:
        excel_file_name = DEBUG_EXCEL
        xml_file_name = DEBUG_XML
        xml_output_file_name = DEBUG_OUTPUT_XML_NAME
    else:
        num_args = len(sys.argv)
        if num_args <= REQUIRED_NUM_ARGS:
            print "Please provide the Excel and XML files."
            return

        excel_file_name = sys.argv[EXCEL_ARG_INDEX]
        xml_file_name = sys.argv[XML_ARG_INDEX]

        # handle optional command-line argument(s)
        if num_args > REQUIRED_NUM_ARGS + 1:
            xml_output_file_name = sys.argv[OUTPUT_XML_ARG_INDEX]

    # read Excel file
    try:
        excel_workbook = load_workbook(excel_file_name)
    except InvalidFileException:
        # handle invalid files (i.e. file doesn't exist)
        extension = excel_file_name.split(".")[-1]
        if extension != OLD_EXCEL_EXTENSION:
            print "Invalid file."
            return
        # handle older versions of Excel (i.e. .xls extension)
        excel_workbook = open_xls_as_xlsx(excel_file_name)

    # check that the survey worksheet exists
    survey_worksheet = excel_workbook.get_sheet_by_name(SURVEY_SHEET_NAME)
    if survey_worksheet is None:
        print "There exists no %s worksheet in %s." % (SURVEY_SHEET_NAME, excel_file_name)
        return

    # read XML file
    xml_file = open(xml_file_name)
    parsed_xml_document = etree.parse(xml_file)

    root = parsed_xml_document.getroot()
    all_vars = root.findall(".//{*}%s" % VAR_TAG)

    # Check same number of rows
    # TODO: Test this
    if len(survey_worksheet.rows) - 1 != len(all_vars):
        print "There aren't the same number of variables in the Excel spreadsheet and the XML file"
        return
    
    first_row = survey_worksheet.rows[0]

    description_column_index = NULL_INDEX
    literal_question_column_index = NULL_INDEX

    for i in range(len(first_row)):
        cell = first_row[i]
        if cell.value == DESCRIPTION_COLUMN_NAME:
            description_column_index = i
        if cell.value == LITERAL_QUESTION_COLUMN_NAME:
            literal_question_column_index = i

    # check if description and question columns exist in excel file
    if description_column_index == NULL_INDEX:
        print "No %s column in %s" % (DESCRIPTION_COLUMN_NAME, excel_file_name)
        return

    if literal_question_column_index == NULL_INDEX:
        print "No %s column in %s" % (LITERAL_QUESTION_COLUMN_NAME, excel_file_name)
        return

    survey_columns = survey_worksheet.columns
    description_columns = survey_columns[description_column_index][1:]
    literal_question_columns = survey_columns[literal_question_column_index][1:]
    
    for c in range(xml_var_num):
        var = all_vars[c]
        # if there is no literal question already
        if var.find("%s%s" % (IGNORE_NAMESPACE, QUESTION_TAG)) is None:
            # create the xml element that represent the question
            question_elt = etree.SubElement(var, QUESTION_TAG)
            literal_question_elt = etree.SubElement(question_elt, LITERAL_QUESTION_TAG)
            literal_question_elt.text = literal_question_columns[c].value

        # check if there is no description already
        if var.find("%s%s" % (IGNORE_NAMESPACE, DESCRIPTION_TAG)) is None:
            # create the xml element that represents the description
            description_elt = etree.SubElement(var, DESCRIPTION_TAG)
            description_elt.text = description_columns[c].value

    # output the resulting "filled in" DDI .xml file
    output = open(xml_output_file_name, "w")
    output.write(etree.tostring(root, pretty_print=True))
    output.close()

if __name__ == '__main__':
    main()